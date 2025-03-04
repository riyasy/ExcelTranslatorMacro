import os
import re
import time

from google.cloud import translate_v2 as translate
from openpyxl import Workbook, load_workbook

os.environ[
    'GOOGLE_APPLICATION_CREDENTIALS'] = 'eighth-epsilon-313100-674cc277361c.json'

translate_client = translate.Client()


def check_if_jap(text: str):
    regex = u'[\u3040-\u30ff\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff\uff66-\uff9f]'
    # regex = u'[\p{Hiragana}\p{Katakana}\p{Han}]+'
    p = re.compile(regex, re.U)
    if p.search(text):
        return True
    else:
        return False


def translate_excel():
    wb = load_workbook('D:\\Translation\Input\in.xlsx')
    for n, sheet in enumerate(wb.worksheets):
        print('Sheet Index:[{}], Title:{}'.format(n, sheet.title))
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                if (cell.value != None):
                    cellValueBefore = cell.value
                    if isinstance(cellValueBefore, str):
                        if cellValueBefore.startswith('=') == False:
                            if check_if_jap(cellValueBefore) == True:
                                cellValueAfter = translate_client.translate(cellValueBefore, target_language='en',
                                                                            source_language='ja',
                                                                            format_='text')
                                # cellValueAfter = cellValueBefore
                                cell.value = cellValueAfter['translatedText']

    wb.save('D:\\Translation\Input\out.xlsx')
    wb.close()


translate_excel()
